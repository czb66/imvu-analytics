"""
数据导出服务 - 支持 CSV/Excel 格式导出
"""

import csv
import io
import logging
from datetime import datetime
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db_context, ProductDataRepository, DatasetRepository
from app.services.analytics import AnalyticsService

logger = logging.getLogger(__name__)


# Excel 表头样式
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def _get_dashboard_summary_data(user_id: int) -> Dict:
    """获取仪表盘汇总数据"""
    with get_db_context() as db:
        repo = ProductDataRepository(db)
        products = repo.get_all(user_id=user_id)
        
        if not products:
            return {
                "summary": {},
                "products": []
            }
        
        # 转换为字典
        product_dicts = []
        for p in products:
            product_dicts.append({
                'product_id': p.product_id,
                'product_name': p.product_name,
                'price': p.price or 0,
                'profit': p.profit or 0,
                'visible': p.visible or 'N',
                'direct_sales': p.direct_sales or 0,
                'indirect_sales': p.indirect_sales or 0,
                'promoted_sales': p.promoted_sales or 0,
                'total_sales': (p.direct_sales or 0) + (p.indirect_sales or 0) + (p.promoted_sales or 0),
                'cart_adds': p.cart_adds or 0,
                'wishlist_adds': p.wishlist_adds or 0,
                'organic_impressions': p.organic_impressions or 0,
                'paid_impressions': p.paid_impressions or 0,
                'conversion_rate': 0
            })
        
        # 计算汇总指标
        analytics = AnalyticsService(product_dicts)
        summary = analytics.get_summary_metrics()
        
        return {
            "summary": summary,
            "products": product_dicts,
            "generated_at": datetime.utcnow().isoformat()
        }


def _get_products_by_dataset(user_id: int, dataset_id: int) -> List[Dict]:
    """获取指定数据集的产品数据"""
    from app.models import Dataset
    with get_db_context() as db:
        # 验证数据集归属
        dataset = db.query(Dataset).filter(Dataset.id == dataset_id, Dataset.user_id == user_id).first()
        if not dataset:
            return []
        
        products = db.query(ProductData).filter(ProductData.dataset_id == dataset_id).all()
        
        if not products:
            return []
        
        product_dicts = []
        for p in products:
            product_dicts.append({
                'product_id': p.product_id,
                'product_name': p.product_name,
                'price': p.price or 0,
                'profit': p.profit or 0,
                'visible': p.visible or 'N',
                'direct_sales': p.direct_sales or 0,
                'indirect_sales': p.indirect_sales or 0,
                'promoted_sales': p.promoted_sales or 0,
                'total_sales': (p.direct_sales or 0) + (p.indirect_sales or 0) + (p.promoted_sales or 0),
                'cart_adds': p.cart_adds or 0,
                'wishlist_adds': p.wishlist_adds or 0,
                'organic_impressions': p.organic_impressions or 0,
                'paid_impressions': p.paid_impressions or 0,
            })
        
        return product_dicts


def export_dashboard_csv(user_id: int) -> io.StringIO:
    """
    导出仪表盘汇总数据为 CSV
    
    Args:
        user_id: 用户ID
    
    Returns:
        CSV 文件内容的 StringIO 对象
    """
    logger.info(f"用户 {user_id} 导出仪表盘数据为 CSV")
    
    data = _get_dashboard_summary_data(user_id)
    products = data.get("products", [])
    summary = data.get("summary", {})
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 写入汇总信息
    writer.writerow(["IMVU Analytics - Dashboard Export"])
    writer.writerow(["Generated At", data.get("generated_at", "")])
    writer.writerow([])
    
    # 写入汇总指标
    writer.writerow(["Summary Metrics"])
    writer.writerow(["Total Products", summary.get("total_products", 0)])
    writer.writerow(["Visible Products", summary.get("visible_products", 0)])
    writer.writerow(["Hidden Products", summary.get("hidden_products", 0)])
    writer.writerow(["Total Sales", round(summary.get("total_sales", 0), 2)])
    writer.writerow(["Total Profit", round(summary.get("total_profit", 0), 2)])
    writer.writerow([])
    
    # 写入产品明细
    if products:
        writer.writerow(["Product Details"])
        headers = [
            "Product ID", "Product Name", "Price", "Profit", "Visible",
            "Direct Sales", "Indirect Sales", "Promoted Sales", "Total Sales",
            "Cart Adds", "Wishlist Adds", "Organic Impressions", "Paid Impressions"
        ]
        writer.writerow(headers)
        
        for p in products:
            writer.writerow([
                p.get('product_id', ''),
                p.get('product_name', ''),
                round(p.get('price', 0), 2),
                round(p.get('profit', 0), 2),
                p.get('visible', 'N'),
                round(p.get('direct_sales', 0), 2),
                round(p.get('indirect_sales', 0), 2),
                round(p.get('promoted_sales', 0), 2),
                round(p.get('total_sales', 0), 2),
                p.get('cart_adds', 0),
                p.get('wishlist_adds', 0),
                p.get('organic_impressions', 0),
                p.get('paid_impressions', 0),
            ])
    
    output.seek(0)
    return output


def export_dashboard_excel(user_id: int) -> io.BytesIO:
    """
    导出仪表盘汇总数据为 Excel（带格式）
    
    Args:
        user_id: 用户ID
    
    Returns:
        Excel 文件内容的 BytesIO 对象
    """
    logger.info(f"用户 {user_id} 导出仪表盘数据为 Excel")
    
    data = _get_dashboard_summary_data(user_id)
    products = data.get("products", [])
    summary = data.get("summary", {})
    
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # 标题
    ws_summary['A1'] = "IMVU Analytics - Dashboard Export"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A2'] = f"Generated At: {data.get('generated_at', '')}"
    
    # 汇总指标
    ws_summary['A4'] = "Summary Metrics"
    ws_summary['A4'].font = Font(bold=True)
    ws_summary['A4'].fill = HEADER_FILL
    ws_summary['A4'].font = HEADER_FONT
    
    summary_data = [
        ["Total Products", summary.get("total_products", 0)],
        ["Visible Products", summary.get("visible_products", 0)],
        ["Hidden Products", summary.get("hidden_products", 0)],
        ["Total Sales", round(summary.get("total_sales", 0), 2)],
        ["Total Profit", round(summary.get("total_profit", 0), 2)],
        ["Direct Sales", round(summary.get("direct_sales", 0), 2)],
        ["Indirect Sales", round(summary.get("indirect_sales", 0), 2)],
        ["Promoted Sales", round(summary.get("promoted_sales", 0), 2)],
    ]
    
    for i, (label, value) in enumerate(summary_data, start=5):
        ws_summary[f'A{i}'] = label
        ws_summary[f'B{i}'] = value
    
    # 调整列宽
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    
    # Sheet 2: Products
    if products:
        ws_products = wb.create_sheet("Products")
        
        # 表头
        headers = [
            "Product ID", "Product Name", "Price", "Profit", "Visible",
            "Direct Sales", "Indirect Sales", "Promoted Sales", "Total Sales",
            "Cart Adds", "Wishlist Adds", "Organic Impressions", "Paid Impressions"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws_products.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        
        # 数据行
        for row_idx, p in enumerate(products, start=2):
            ws_products.cell(row=row_idx, column=1, value=p.get('product_id', '')).border = THIN_BORDER
            ws_products.cell(row=row_idx, column=2, value=p.get('product_name', '')).border = THIN_BORDER
            
            # 数值列
            numeric_cols = [3, 4, 6, 7, 8, 9, 10, 11, 12, 13]
            values = [
                round(p.get('price', 0), 2),
                round(p.get('profit', 0), 2),
                round(p.get('direct_sales', 0), 2),
                round(p.get('indirect_sales', 0), 2),
                round(p.get('promoted_sales', 0), 2),
                round(p.get('total_sales', 0), 2),
                p.get('cart_adds', 0),
                p.get('wishlist_adds', 0),
                p.get('organic_impressions', 0),
                p.get('paid_impressions', 0),
            ]
            for col, val in zip(numeric_cols, values):
                cell = ws_products.cell(row=row_idx, column=col, value=val)
                cell.border = THIN_BORDER
                if col in [3, 4, 6, 7, 8, 9]:  # 货币/金额列
                    cell.number_format = '#,##0.00'
            
            ws_products.cell(row=row_idx, column=5, value=p.get('visible', 'N')).border = THIN_BORDER
        
        # 调整列宽
        ws_products.column_dimensions['A'].width = 15
        ws_products.column_dimensions['B'].width = 40
        for col in range(3, 14):
            ws_products.column_dimensions[get_column_letter(col)].width = 12
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_products_csv(user_id: int, dataset_id: int) -> io.StringIO:
    """
    导出产品明细为 CSV
    
    Args:
        user_id: 用户ID
        dataset_id: 数据集ID
    
    Returns:
        CSV 文件内容的 StringIO 对象
    """
    logger.info(f"用户 {user_id} 导出数据集 {dataset_id} 产品数据为 CSV")
    
    products = _get_products_by_dataset(user_id, dataset_id)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 写入标题
    writer.writerow(["IMVU Analytics - Products Export"])
    writer.writerow(["Dataset ID", dataset_id])
    writer.writerow(["Exported At", datetime.utcnow().isoformat()])
    writer.writerow([])
    
    if products:
        headers = [
            "Product ID", "Product Name", "Price", "Profit", "Visible",
            "Direct Sales", "Indirect Sales", "Promoted Sales", "Total Sales",
            "Cart Adds", "Wishlist Adds", "Organic Impressions", "Paid Impressions"
        ]
        writer.writerow(headers)
        
        for p in products:
            writer.writerow([
                p.get('product_id', ''),
                p.get('product_name', ''),
                round(p.get('price', 0), 2),
                round(p.get('profit', 0), 2),
                p.get('visible', 'N'),
                round(p.get('direct_sales', 0), 2),
                round(p.get('indirect_sales', 0), 2),
                round(p.get('promoted_sales', 0), 2),
                round(p.get('total_sales', 0), 2),
                p.get('cart_adds', 0),
                p.get('wishlist_adds', 0),
                p.get('organic_impressions', 0),
                p.get('paid_impressions', 0),
            ])
    else:
        writer.writerow(["No products found in this dataset."])
    
    output.seek(0)
    return output


def export_products_excel(user_id: int, dataset_id: int) -> io.BytesIO:
    """
    导出产品明细为 Excel（带格式）
    
    Args:
        user_id: 用户ID
        dataset_id: 数据集ID
    
    Returns:
        Excel 文件内容的 BytesIO 对象
    """
    logger.info(f"用户 {user_id} 导出数据集 {dataset_id} 产品数据为 Excel")
    
    products = _get_products_by_dataset(user_id, dataset_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # 标题
    ws['A1'] = f"IMVU Analytics - Products Export (Dataset {dataset_id})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Exported At: {datetime.utcnow().isoformat()}"
    
    if products:
        # 表头
        headers = [
            "Product ID", "Product Name", "Price", "Profit", "Visible",
            "Direct Sales", "Indirect Sales", "Promoted Sales", "Total Sales",
            "Cart Adds", "Wishlist Adds", "Organic Impressions", "Paid Impressions"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        
        # 数据行
        for row_idx, p in enumerate(products, start=5):
            ws.cell(row=row_idx, column=1, value=p.get('product_id', '')).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=p.get('product_name', '')).border = THIN_BORDER
            
            numeric_cols = [3, 4, 6, 7, 8, 9, 10, 11, 12, 13]
            values = [
                round(p.get('price', 0), 2),
                round(p.get('profit', 0), 2),
                round(p.get('direct_sales', 0), 2),
                round(p.get('indirect_sales', 0), 2),
                round(p.get('promoted_sales', 0), 2),
                round(p.get('total_sales', 0), 2),
                p.get('cart_adds', 0),
                p.get('wishlist_adds', 0),
                p.get('organic_impressions', 0),
                p.get('paid_impressions', 0),
            ]
            for col, val in zip(numeric_cols, values):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = THIN_BORDER
                if col in [3, 4, 6, 7, 8, 9]:
                    cell.number_format = '#,##0.00'
            
            ws.cell(row=row_idx, column=5, value=p.get('visible', 'N')).border = THIN_BORDER
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        for col in range(3, 14):
            ws.column_dimensions[get_column_letter(col)].width = 12
    else:
        ws['A4'] = "No products found in this dataset."
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
